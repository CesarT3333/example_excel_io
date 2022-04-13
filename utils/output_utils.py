import datetime

from openpyxl import Workbook
from datetime import datetime
from entity.configuration import Configuration

wb = Workbook()


def define_file_name(configuration: Configuration):
    # FunctionalCatalog{environment}{Date}
    now = datetime.now()
    file_name: str = 'FunctionalCatalog' + '' + configuration.env.lower() + '_' + now.strftime("%Y%m%d").title()

    file_output_path = 'resources/output/'
    file_format = '.xlsx'

    configuration.output_file_name = file_output_path + file_name + file_format
    wb.save(configuration.output_file_name)


def create_sheets(configuration: Configuration):
    for sheetName in configuration.sheets():
        wb.create_sheet(sheetName)

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    wb.save(configuration.output_file_name)


def show_importation_success_log(configuration: Configuration):
    print('File:           ', configuration.input_file_path, 'successfully importing')
    print('File generated: ', configuration.output_file_name)
